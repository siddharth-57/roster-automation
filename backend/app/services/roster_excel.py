from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.roster import Roster
from app.models.roster_assignment import RosterAssignment
from app.models.team_member import TeamMember


# ------------------------------------------------------
# EXCEL HELPERS
# ------------------------------------------------------

SHIFT_CODES = {"A", "B", "C", "G", "L", "W", "H"}


def get_days_in_month(year: int, month: int) -> int:
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return (next_month - date(year, month, 1)).days


def get_roster_name(
    year: int,
    month: int,
    group_number: str,
) -> str:
    return (
        f"{year}-{month:02d}-Group-{group_number}"
    )


# ------------------------------------------------------
# EXPORT ROSTER TO EXCEL
# ------------------------------------------------------

def create_roster_excel(
    db: Session,
    roster: Roster,
) -> BytesIO:
    """
    Create an Excel workbook from an existing roster.

    The workbook follows the reference roster format:
    - Row 1 contains group, serial number, member name,
      and dates.
    - Row 2 contains weekdays.
    - Following rows contain one member per row.
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Roster"

    # --------------------------------------------------
    # Get active team members.
    # --------------------------------------------------

    # --------------------------------------------------
    # Get employees who actually have assignments in
    # this roster.
    #
    # We intentionally do NOT filter by active=True.
    # A deactivated employee may still belong to an
    # old/historical roster.
    # --------------------------------------------------

    assignments = db.scalars(
        select(RosterAssignment)
        .where(
            RosterAssignment.roster_id
            == roster.roster_id
        )
    ).all()


    employee_ids = {
        assignment.employee_id
        for assignment in assignments
    }


    # --------------------------------------------------
    # Get those employees from team_members and order
    # them according to their current display_order.
    # --------------------------------------------------

    members = db.scalars(
        select(TeamMember)
        .where(
            TeamMember.employee_id.in_(employee_ids)
        )
        .order_by(TeamMember.display_order)
    ).all()


    # --------------------------------------------------
    # Get all roster assignments.
    # --------------------------------------------------

    assignment_lookup = {}

    for assignment in assignments:
        assignment_lookup[
            (
                assignment.employee_id,
                assignment.date.day,
            )
        ] = assignment.shift

    days_in_month = get_days_in_month(
        roster.year,
        roster.month,
    )

    # --------------------------------------------------
    # Row 1
    # --------------------------------------------------

    worksheet.cell(
        row=1,
        column=1,
        value=f"Group = {roster.group_number}",
    )

    worksheet.cell(
        row=1,
        column=2,
        value="S.No",
    )

    worksheet.cell(
        row=1,
        column=3,
        value="MOCC Members",
    )

    # --------------------------------------------------
    # Row 1 dates
    # --------------------------------------------------

    for day in range(1, days_in_month + 1):
        column = day + 3

        worksheet.cell(
            row=1,
            column=column,
            value=f"{day:02d}-{date(roster.year, roster.month, day).strftime('%b')}",
        )

    # --------------------------------------------------
    # Row 2 weekdays
    # --------------------------------------------------

    for day in range(1, days_in_month + 1):
        column = day + 3

        worksheet.cell(
            row=2,
            column=column,
            value=date(
                roster.year,
                roster.month,
                day,
            ).strftime("%a"),
        )

    # --------------------------------------------------
    # Member rows
    # --------------------------------------------------

    for row_index, member in enumerate(
        members,
        start=3,
    ):
        serial_number = row_index - 2
    
        worksheet.cell(
            row=row_index,
            column=2,
            value=serial_number,
        )

        worksheet.cell(
            row=row_index,
            column=3,
            value=member.name,
        )

        for day in range(1, days_in_month + 1):
            column = day + 3

            shift = assignment_lookup.get(
                (
                    member.employee_id,
                    day,
                )
            )

            worksheet.cell(
                row=row_index,
                column=column,
                value=shift,
            )

    # --------------------------------------------------
    # Basic formatting
    # --------------------------------------------------

    bold_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.font = bold_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for cell in worksheet[2]:
        cell.font = bold_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row in worksheet.iter_rows(
        min_row=3,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    # --------------------------------------------------
    # Column widths
    # --------------------------------------------------

    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 8
    worksheet.column_dimensions["C"].width = 22

    for column in range(4, days_in_month + 4):
        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = 10

    # --------------------------------------------------
    # Freeze member/name headers.
    # --------------------------------------------------

    worksheet.freeze_panes = "D3"

    # --------------------------------------------------
    # Return workbook as an in-memory file.
    # --------------------------------------------------

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# ------------------------------------------------------
# IMPORT ROSTER FROM EXCEL
# ------------------------------------------------------

def read_roster_excel(
    db: Session,
    file_contents: bytes,
    year: int,
    month: int,
    group_number: str,
) -> dict[str, dict[int, str]]:
    """
    Read a roster Excel file and convert it into the
    internal roster representation:

        {
            "E001": {
                1: "A",
                2: "W",
                ...
            }
        }

    The employee name in the Excel file is matched against
    the team member table.
    """

    workbook = load_workbook(
        BytesIO(file_contents),
        data_only=True,
    )

    # worksheet = workbook.active
    worksheet = workbook.worksheets[0]  

    days_in_month = get_days_in_month(
        year,
        month,
    )

    # --------------------------------------------------
    # Validate the group number.
    # --------------------------------------------------

    group_value = worksheet.cell(
        row=1,
        column=1,
    ).value

    expected_group = (
        f"Group = {group_number}"
    )

    if str(group_value).strip() != expected_group:
        raise ValueError(
            f"Excel group does not match "
            f"the selected group number. "
            f"Expected '{expected_group}'."
        )

    # --------------------------------------------------
    # Load team members.
    # --------------------------------------------------

    members = db.scalars(
        select(TeamMember)
    ).all()

    member_by_name = {
        member.name.strip().lower(): member
        for member in members
    }

    roster = {}

    # --------------------------------------------------
    # Read member rows.
    # --------------------------------------------------

    for row in range(
        3,
        worksheet.max_row + 1,
    ):
        name_value = worksheet.cell(
            row=row,
            column=3,
        ).value

        if not name_value:
            continue

        name = str(name_value).strip()

        member = member_by_name.get(
            name.lower()
        )

        if not member:
            raise ValueError(
                f"Employee '{name}' from the "
                f"Excel file does not exist "
                f"in the team members table."
            )

        employee_id = member.employee_id

        roster[employee_id] = {}

        for day in range(
            1,
            days_in_month + 1,
        ):
            column = day + 3

            value = worksheet.cell(
                row=row,
                column=column,
            ).value

            if value is None:
                continue

            shift = str(value).strip().upper()

            if shift == "":
                continue

            if shift not in SHIFT_CODES:
                raise ValueError(
                    f"Invalid shift '{shift}' "
                    f"for employee '{name}' "
                    f"on day {day}."
                )

            roster[employee_id][day] = shift

    return roster